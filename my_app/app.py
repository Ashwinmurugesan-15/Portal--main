from flask import Flask, jsonify, request, render_template, redirect, url_for, session, send_from_directory
from flask_cors import CORS
import os, re, tempfile, logging, time
import openpyxl
from openpyxl.cell.cell import MergedCell
from datetime import datetime
import random
import json
from collections import defaultdict, Counter
import secrets
import sqlite3
import hashlib
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pypdf import PdfReader
import requests

from jinja2 import FileSystemLoader, ChoiceLoader
from config import DevelopmentConfig, ProductionConfig

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

env = os.getenv("APP_ENV", "development").lower()
if env == "production":
    app_config = ProductionConfig()
else:
    app_config = DevelopmentConfig()

base_dir = os.path.dirname(os.path.abspath(__file__))
templates_dir = os.path.join(base_dir, "templates")
static_dir = os.path.join(base_dir, "static")

# PDF Upload Configuration
UPLOAD_FOLDER = os.path.join(base_dir, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__, static_folder=static_dir, template_folder=templates_dir)
app.config.from_object(app_config)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = app.config.get("SECRET_KEY") or None

if not app.secret_key:
    if env == "production":
        raise RuntimeError("SECRET_KEY must be set in production")
    dev_secret = os.getenv("DEV_SECRET_KEY")
    if dev_secret:
        app.secret_key = dev_secret
    else:
        generated_secret = secrets.token_hex(32)
        app.secret_key = generated_secret
        app.config["SECRET_KEY"] = generated_secret

logger.info("Flask SECRET_KEY configured for env '%s': %s", env, "set" if app.secret_key else "missing")
api_allowed_origins = app.config.get("API_ALLOWED_ORIGINS", "*")
if api_allowed_origins and api_allowed_origins != "*":
    origins_list = [o.strip() for o in str(api_allowed_origins).split(",") if o.strip()]
    CORS(app, resources={r"/api/*": {"origins": origins_list}}, supports_credentials=True)
else:
    CORS(app, supports_credentials=True)

app.jinja_env.loader = ChoiceLoader([
    FileSystemLoader(templates_dir),
    FileSystemLoader(app.root_path)
])

EXCEL_FILE = app.config["EXCEL_FILE"]
SHEET_NAME = app.config["SHEET_NAME"]
USER_DB = app.config["USER_DB"]
DATABASE = app.config["DATABASE"]
EMAIL_CONFIG = app.config["EMAIL_CONFIG"]
ADMIN_USERNAME = app.config["ADMIN_USERNAME"]
ADMIN_PASSWORD = app.config["ADMIN_PASSWORD"]

if env == "production":
    if not app.secret_key:
        raise RuntimeError("SECRET_KEY must be set in production")
    if not ADMIN_USERNAME or not ADMIN_PASSWORD:
        raise RuntimeError("ADMIN_USERNAME and ADMIN_PASSWORD must be set in production")
    if not EMAIL_CONFIG.get("SENDER_EMAIL") or not EMAIL_CONFIG.get("SENDER_PASSWORD"):
        raise RuntimeError("Email sender credentials must be set in production")


# ============================================
# Token Manager for Guhatek API Integration
# ============================================
class TokenManager:
    """Manages API token with automatic refresh every 10 minutes"""
    
    def __init__(self):
        self.token = None
        self.token_expiry = None
        self.api_base_url = "https://api-dev.guhatek.org"
        self.api_key = "guhatek-job-applicant"
    
    def get_token(self):
        """Get current token or fetch new one if expired"""
        current_time = time.time()
        
        # Check if token exists and is still valid (refresh 30 seconds before expiry)
        if self.token and self.token_expiry and current_time < (self.token_expiry - 30):
            logger.info("Using cached token")
            return self.token
        
        # Fetch new token
        logger.info("Fetching new token from Guhatek API")
        try:
            response = requests.get(
                f"{self.api_base_url}/api/token",
                headers={"x-api-key": self.api_key},
                timeout=10
            )
            response.raise_for_status()
            
            data = response.json()
            self.token = data.get("token")
            
            # Set expiry to 10 minutes from now (600 seconds)
            self.token_expiry = current_time + 600
            
            logger.info(f"New token fetched, expires in 10 minutes")
            return self.token
            
        except Exception as e:
            logger.error(f"Error fetching token: {str(e)}")
            raise

# Initialize global token manager
token_manager = TokenManager()


@app.context_processor
def inject_app_config():
    return {"APP_CONFIG": app.config}

# Create sample Excel file if it doesn't exist
def create_sample_excel():
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)

    wb = openpyxl.Workbook()
    sheet = wb.active
    if sheet is not None:
        sheet.title = SHEET_NAME
    
    # Define headers
    headers = [
        'Date', 'Name', 'Email ID', 'Contact Number', 'LinkedIn Profile', 'Resume',
        'Interested Position', 'Current Role', 'Current Organization', 'Total Years of Experience',
        'Current Location', 'Location Preference', 'Current CTC per Annum', 'Expected CTC per Annum',
        'Notice Period', 'In Notice', 'Immediate Joiner', 'Offers in Hand', 'Offered CTC',
        'Certifications', 'Referred By',
        'Interview Status', 'Application Status',
        'Initial Screening', 'Round 1 D and T', 'Round 1 Remarks', 'Round 2 D and T', 'Round 2 Remarks',
        'Offered Position', 'Joining Date', 'Reject Mail Sent', 'Remarks', 'Screened By'
    ]
    
    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        if sheet is not None:
            cell = sheet.cell(row=1, column=col_num)
            # Check if cell is not a merged cell before assigning value
            if cell is not None and not isinstance(cell, MergedCell):
                cell.value = header
        
        # Sample data
        sample_data = [
            {
                'Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Name': 'John Doe',
                'Email ID': 'john.doe@example.com',
                'Contact Number': '9876543210',
                'Interested Position': 'Software Developer',
                'Current Role': 'Junior Developer',
                'Current Organization': 'Tech Solutions Inc.',
                'Current Location': 'Bangalore',
                'Current CTC per Annum': '800000',
                'Expected CTC per Annum': '1200000',
                'Total Years of Experience': '2-3 years',
                'Notice Period': '30 days',
                'In Notice': 'Yes',
                'Immediate Joiner': 'No',
                'Offers in Hand': 'No',
                'Offered CTC': '',
                'Location Preference': 'Bangalore',
                'Certifications': 'AWS Certified Developer',
                'Resume': 'https://example.com/resume/johndoe',
                'LinkedIn Profile': 'https://linkedin.com/in/johndoe',
                'Comments': 'Good communication skills',
                'Referred By': 'Employee Referral',
                'Interview Status': 'Scheduled',
                'Application Status': 'In Process',
                'Remarks': 'Promising candidate',
                'Reject Mail Sent': 'No',
                'Initial Screening': 'Candidate performed well in initial screening.',
                'Round 1 Remarks': 'Strong technical skills demonstrated in Round 1.',
                'Round 2 Remarks': 'Good problem-solving approach in Round 2.',
               
                'Reference': 'Jane Smith'
            },
            {
                'Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Name': 'Jane Smith',
                'Email ID': 'jane.smith@example.com',
                'Contact Number': '8765432109',
                'Interested Position': 'Data Scientist',
                'Current Role': 'Data Analyst',
                'Current Organization': 'Data Insights Ltd.',
                'Current Location': 'Hyderabad',
                'Current CTC per Annum': '1000000',
                'Expected CTC per Annum': '1500000',
                'Total Years of Experience': '3-5 years',
                'Notice Period': '60 days',
                'In Notice': 'No',
                'Immediate Joiner': 'No',
                'Offers in Hand': 'Yes',
                'Offered CTC': '1400000',
                'Location Preference': 'Remote',
                'Certifications': 'Google Data Analytics',
                'Resume': 'https://example.com/resume/janesmith',
                'LinkedIn Profile': 'https://linkedin.com/in/janesmith',
                'Comments': 'Strong analytical skills',
                'Referred By': 'Job Portal',
                'Interview Status': 'Selected',
                'Application Status': 'Offer Made',
                'Remarks': 'Top candidate',
                'Reject Mail Sent': 'No',
                'Initial Remarks': '',
                'Round 1 Remarks': '',
                'Round 2 Remarks': '',
                
                'Reference': 'Robert Johnson'
            },
            {
                'Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Email ID': 'sam.wilson@example.com',
                'Contact Number': '7654321098',
                'Interested Position': 'UI/UX Designer',
                'Current Role': 'Graphic Designer',
                'Current Organization': 'Creative Designs',
                'Current Location': 'Chennai',
                'Current CTC per Annum': '700000',
                'Expected CTC per Annum': '1000000',
                'Total Years of Experience': '1-2 years',
                'Notice Period': '15 days',
                'In Notice': 'Yes',
                'Immediate Joiner': 'Yes',
                'Offers in Hand': 'No',
                'Offered CTC': '',
                'Location Preference': 'Chennai',
                'Certifications': 'Adobe Certified Expert',
                'Resume': 'https://example.com/resume/samwilson',
                'LinkedIn Profile': 'https://linkedin.com/in/samwilson',
                'Comments': 'Creative portfolio',
                'Referred By': 'Campus Recruitment',
                'Interview Status': 'Rejected',
                'Application Status': 'Rejected',
                'Remarks': 'Not enough experience',
                'Reject Mail Sent': 'Yes',
                'Initial Remarks': '',
                'Round 1 Remarks': '',
                'Round 2 Remarks': '',
                
                'Reference': 'Emily Davis'
            }
        ]
        
        # Add sample data
        for row_num, data in enumerate(sample_data, 2):
            for col_num, header in enumerate(headers, 1):
                if sheet is not None:
                    cell = sheet.cell(row=row_num, column=col_num)
                    # Check if cell is not a merged cell before assigning value
                    if cell is not None and not isinstance(cell, MergedCell):
                        cell.value = data.get(header, '')
        
        # Save the workbook
        wb.save(EXCEL_FILE)
        wb.close()
        print(f"Created sample Excel file: {EXCEL_FILE}")

# Load data from Excel
def load_data():
    try:
        if not os.path.exists(EXCEL_FILE):
            create_sample_excel()
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb[SHEET_NAME]
        
        # Get headers from the first row
        headers = [cell.value for cell in sheet[1] if cell is not None]
        
        # Get data from the remaining rows
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i, value in enumerate(row):
                # Convert datetime objects to string
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                header = headers[i]
                # Migrate old "Initial Remarks" to "Initial Screening"
                if header == 'Initial Remarks':
                    header = 'Initial Screening'
                row_data[header] = str(value) if value is not None else ''
            data.append(row_data)
        
        return data
    except Exception as e:
        print(f"Error loading data from Excel: {e}")
        # Optionally, re-raise the exception or return an empty list/error indicator
        return []

# Save data to Excel
def save_data(data):
    try:
        if not os.path.exists(EXCEL_FILE):
            create_sample_excel()

        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb[SHEET_NAME]

        # Get current headers
        headers = [cell.value for cell in sheet[1] if cell is not None and cell.value]
        
        # Desired field order (keep 'Date' at the beginning)
        desired_fields = [
            'Name', 'Email ID', 'Contact Number', 'LinkedIn Profile', 'Resume',
            'Interested Position', 'Current Role', 'Current Organization', 'Total Years of Experience',
            'Current Location', 'Location Preference', 'Current CTC per Annum', 'Expected CTC per Annum',
            'Notice Period', 'In Notice', 'Immediate Joiner', 'Offers in Hand', 'Offered CTC',
            'Certifications', 'Referred By',
            'Interview Status', 'Application Status','Remarks',
            'Initial Screening', 'Round 1 D and T', 'Round 1 Remarks', 'Round 2 D and T', 'Round 2 Remarks',
            'Offered Position', 'Joining Date', 'Reject Mail Sent', 'Screened By',
        ]
        
        # Build ordered headers: Date + desired fields present + any remaining headers
        ordered_headers = []
        if 'Date' in headers:
            ordered_headers.append('Date')
        ordered_headers.extend([h for h in desired_fields if h in headers])
        # Include any headers not in desired list (e.g., 'Reference')
        ordered_headers.extend([h for h in headers if h not in ordered_headers])
        
        # If there are desired fields missing from headers, append them so they are created
        ordered_headers.extend([h for h in desired_fields if h not in ordered_headers])
        
        # Rewrite headers in desired order
        for col_num, header in enumerate(ordered_headers, 1):
            if sheet is not None:
                cell = sheet.cell(row=1, column=col_num)
                if cell is not None and not isinstance(cell, MergedCell):
                    cell.value = header
        
        # Clear existing data (except headers)
        for row in range(sheet.max_row, 1, -1):
            sheet.delete_rows(row)
        
        # Add updated data
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(ordered_headers, 1):
                # Migrate old "Initial Remarks" to "Initial Screening"
                if header == 'Initial Screening':
                    value = row_data.get('Initial Screening') or row_data.get('Initial Remarks', '')
                else:
                    value = row_data.get(header, '')
                # Convert value to string, handle None
                if value is None:
                    value = ''
                else:
                    value = str(value)
                if sheet is not None:
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell is not None and not isinstance(cell, MergedCell):
                        cell.value = value
        
        # Save and close the workbook
        wb.save(EXCEL_FILE)
        wb.close()
        print(f"Data saved to Excel: {len(data)} records")
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in save_data: {error_trace}")
        raise

# Initialize user database
def init_user_db():
    """Initialize the user database with admin user"""
    os.makedirs(os.path.dirname(USER_DB), exist_ok=True)
    conn = sqlite3.connect(USER_DB)
    cursor = conn.cursor()
    
    # Create users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Check if admin user exists
    cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', (ADMIN_USERNAME,))
    if cursor.fetchone()[0] == 0:
        # Create default admin user
        password_hash = hashlib.sha256(ADMIN_PASSWORD.encode()).hexdigest()
        cursor.execute('''
            INSERT INTO users (username, password_hash, is_admin)
            VALUES (?, ?, 1)
        ''', (ADMIN_USERNAME, password_hash))
        conn.commit()
    
    conn.close()

# Hash password
def hash_password(password):
    """Hash a password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

# Verify password
def verify_password(password, password_hash):
    """Verify a password against its hash"""
    return hash_password(password) == password_hash

# Check if user is admin
def is_admin():
    """Check if the current user is an admin"""
    return session.get('is_admin', False)


# Send rejection email
def send_rejection_email(candidate_name, candidate_email, position):
    """Send a professional rejection email to the candidate"""
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['From'] = f"{EMAIL_CONFIG['SENDER_NAME']} <{EMAIL_CONFIG['SENDER_EMAIL']}>"
        msg['To'] = candidate_email
        msg['Subject'] = f"Application Update - {position}"
        
        # HTML email body
        html_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h2 style="color: #2c3e50;">Application Status Update</h2>
                    
                    <p>Dear {candidate_name},</p>
                    
                    <p>Thank you for your interest in the <strong>{position}</strong> position at our organization.</p>
                    
                    <p>After careful consideration, we regret to inform you that we have decided to move forward with other candidates.</p>
                    
                    <p>We appreciate your time and encourage you to apply for future openings.</p>
                    
                    <p>Best regards,<br>
                    <strong>{EMAIL_CONFIG['SENDER_NAME']}</strong></p>
                    
                    <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
                    <p style="font-size: 12px; color: #666;">
                        This is an automated message.
                    </p>
                </div>
            </body>
        </html>
        """
        
        # Plain text version
        text_body = f"""
Dear {candidate_name},

Thank you for your interest in the {position} position.

After careful consideration, we regret to inform you that we have decided to move forward with other candidates.

We appreciate your time and encourage you to apply for future openings.

Best regards,
{EMAIL_CONFIG['SENDER_NAME']}
        """
        
        # Attach both versions
        part1 = MIMEText(text_body, 'plain')
        part2 = MIMEText(html_body, 'html')
        msg.attach(part1)
        msg.attach(part2)
        
        # Send email
        with smtplib.SMTP(EMAIL_CONFIG['SMTP_SERVER'], EMAIL_CONFIG['SMTP_PORT']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['SENDER_EMAIL'], EMAIL_CONFIG['SENDER_PASSWORD'])
            server.send_message(msg)
        
        print(f"✅ Rejection email sent to {candidate_email}")
        return True, "Email sent successfully"
    
    except Exception as e:
        print(f"❌ Error sending email: {str(e)}")
        return False, str(e)

# Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Check against database
        conn = sqlite3.connect(USER_DB)
        cursor = conn.cursor()
        cursor.execute('SELECT password_hash, is_admin FROM users WHERE username = ?', (username,))
        user = cursor.fetchone()
        conn.close()
        
        if user and verify_password(password, user[0]):
            session['logged_in'] = True
            session['username'] = username
            session['is_admin'] = bool(user[1])
            return redirect(url_for('index'))
        else:
            error = 'Invalid credentials. Please try again.'
    
    return render_template('login.html', error=error)

# Logout route
@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    session.pop('username', None)
    session.pop('is_admin', None)
    return redirect(url_for('login'))

# Check if user is logged in
def login_required(f):
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            # For API requests, return JSON error instead of redirect
            if request.path.startswith('/api/'):
                return jsonify({"status": "error", "message": "Authentication required. Please log in."}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Check if user is admin (decorator)
def admin_required(f):
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        if not session.get('is_admin'):
            return jsonify({"status": "error", "message": "Admin access required"}), 403
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
@login_required
def index():
    return render_template('index.html', is_admin=session.get('is_admin', False))

@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/api/data', methods=['GET'])
@login_required
def get_data():
    """Fetch data from Guhatek API (replaces Excel as primary source)"""
    try:
        logger.info("=== Fetching applicants from Guhatek API ===")
        
        # Get token from token manager
        token = token_manager.get_token()
        
        # Call applications API
        response = requests.get(
            f"{token_manager.api_base_url}/api/applications",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            },
            timeout=15
        )
        response.raise_for_status()
        
        api_data = response.json()
        raw_applicants = api_data.get("data", [])
        
        logger.info(f"Received {len(raw_applicants)} applicants from API")
        
        # Filter and transform data - only include complete records
        valid_applicants = []
        for applicant in raw_applicants:
            # Check if critical fields have values
            has_name = applicant.get("full_name") and str(applicant.get("full_name")).strip()
            has_email = applicant.get("email") and str(applicant.get("email")).strip()
            has_contact = applicant.get("contact_number") and str(applicant.get("contact_number")).strip()
            
            if has_name and has_email and has_contact:
                # Transform API response to match frontend format
                transformed = {
                    "Date": applicant.get("submitted_at", ""),
                    "Name": applicant.get("full_name", ""),
                    "Email ID": applicant.get("email", ""),
                    "Contact Number": applicant.get("contact_number", ""),
                    "LinkedIn Profile": applicant.get("linkedin_profile", ""),
                    "Resume": applicant.get("resume_url", ""),
                    "Interested Position": applicant.get("interested_position", ""),
                    "Current Role": applicant.get("currentrole", ""),
                    "Current Organization": applicant.get("current_organisation", ""),
                    "Total Years of Experience": str(applicant.get("total_experience", "")) if applicant.get("total_experience") else "",
                    "Current Location": applicant.get("current_location", ""),
                    "Location Preference": applicant.get("location_preference", ""),
                    "Current CTC per Annum": str(applicant.get("current_ctc", "")) if applicant.get("current_ctc") else "",
                    "Expected CTC per Annum": str(applicant.get("expected_ctc", "")) if applicant.get("expected_ctc") else "",
                    "Notice Period": str(applicant.get("notice_period", "")) if applicant.get("notice_period") else "",
                    "In Notice": "Yes" if applicant.get("currently_noticeperiod") else "No",
                    "Immediate Joiner": "Yes" if applicant.get("immediate_joiner") else "No",
                    "Offers in Hand": "Yes" if applicant.get("other_offer_in_hand") else "No",
                    "Offered CTC": str(applicant.get("offered_ctc", "")) if applicant.get("offered_ctc") else "",
                    "Certifications": applicant.get("certifications", "") or "",
                    "Referred By": applicant.get("referred_by", "") or "",
                    "Interview Status": "",
                    "Application Status": "",
                    "Initial Screening": "",
                    "Round 1 D and T": "",
                    "Round 1 Remarks": "",
                    "Round 2 D and T": "",
                    "Round 2 Remarks": "",
                    "Offered Position": "",
                    "Joining Date": "",
                    "Reject Mail Sent": "No",
                    "Remarks": applicant.get("additional_info", "") or "",
                    "Screened By": ""
                }
                valid_applicants.append(transformed)
        
        logger.info(f"Returning {len(valid_applicants)} valid applicants")
        
        is_admin_user = is_admin()
        return jsonify({"data": valid_applicants, "is_admin": is_admin_user})
        
    except requests.exceptions.Timeout:
        logger.error("Guhatek API timeout - falling back to Excel")
        # Fallback to Excel if API fails
        data = load_data()
        return jsonify({"data": data, "is_admin": is_admin()})
    except Exception as e:
        logger.error(f"Error fetching from API: {str(e)} - falling back to Excel")
        # Fallback to Excel if API fails
        data = load_data()
        return jsonify({"data": data, "is_admin": is_admin()})

@app.route('/api/data', methods=['POST'])
@login_required
def add_data():
    try:
        # Handle multipart/form-data (for file uploads) or application/json
        if request.is_json:
            new_data = request.json
        else:
            new_data = request.form.to_dict()
            if 'Resume' in request.files:
                file = request.files['Resume']
                if file and file.filename:
                    # Create a safe, timestamped filename
                    filename = f"{int(time.time())}_{file.filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    new_data['Resume'] = filename
        
        data = load_data()
        data.append(new_data)
        save_data(data)
        return jsonify({"status": "success", "message": "Data added successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/data/<int:index>', methods=['PUT'])
@login_required
def update_data(index):
    try:
        # Handle multipart/form-data (for file uploads) or application/json
        if request.is_json:
            update_payload = request.json
        else:
            update_payload = request.form.to_dict()
            if 'Resume' in request.files:
                file = request.files['Resume']
                if file and file.filename:
                    # Create a safe, timestamped filename
                    filename = f"{int(time.time())}_{file.filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    update_payload['Resume'] = filename
        
        data = load_data()
        
        # Check if index is valid
        if 0 <= index < len(data):
            # Track if screening/remarks fields are being updated
            screening_fields = ['Initial Screening', 'Remarks', 'Round 1 Remarks', 'Round 2 Remarks']
            is_screening_update = any(
                key in screening_fields and update_payload.get(key) 
                for key in update_payload.keys()
            )
            
            # Update the data at the specified index
            for key, value in update_payload.items():
                # Convert specific fields to appropriate types if necessary
                if key in ['Current CTC per Annum', 'Expected CTC per Annum', 'Offered CTC']:
                    try:
                        data[index][key] = int(value) if value else ''
                    except (ValueError, TypeError):
                        data[index][key] = value  # Keep original if conversion fails
                else:
                    # Ensure all values are strings or None
                    data[index][key] = str(value) if value is not None else ''
            
            # If screening/remarks fields were updated, track the user
            if is_screening_update:
                current_user = session.get('username', 'Unknown')
                data[index]['Screened By'] = current_user
            
            save_data(data)
            return jsonify({"status": "success", "message": "Data updated successfully"})
        else:
            return jsonify({"status": "error", "message": f"No record found at index {index}"}), 404
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error updating data: {error_trace}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/data/<int:index>', methods=['DELETE'])
@login_required
def delete_data(index):
    try:
        data = load_data()
        
        # Check if index is valid
        if 0 <= index < len(data):
            # Delete the data at the specified index
            del data[index]
            save_data(data)
            return jsonify({"status": "success", "message": "Data deleted successfully"})
        else:
            return jsonify({"status": "error", "message": f"No record found at index {index}"}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/applicants', methods=['GET'])
@login_required
def get_applicants_from_api():
    """Fetch applicants from Guhatek API and filter out null records"""
    try:
        logger.info("=== Starting API applicants fetch ===")
        
        # Get token from token manager
        logger.info("Getting token from token manager...")
        token = token_manager.get_token()
        logger.info(f"Token obtained: {token[:20]}..." if token else "No token!")
        
        # Call applications API
        logger.info("Fetching applicants from Guhatek API")
        response = requests.get(
            f"{token_manager.api_base_url}/api/applications",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            },
            timeout=15  # Increased timeout for slower dev cluster
        )
        logger.info(f"API response status: {response.status_code}")
        response.raise_for_status()
        
        data = response.json()
        raw_applicants = data.get("data", [])
        
        logger.info(f"Received {len(raw_applicants)} applicants from API")
        
        # Filter out records with null values for critical fields
        # Only include records that have AT LEAST: full_name, email, and contact_number
        valid_applicants = []
        for applicant in raw_applicants:
            # Check if critical fields have values (not null and not empty string)
            has_name = applicant.get("full_name") and str(applicant.get("full_name")).strip()
            has_email = applicant.get("email") and str(applicant.get("email")).strip()
            has_contact = applicant.get("contact_number") and str(applicant.get("contact_number")).strip()
            
            # Only include if all critical fields have values
            if has_name and has_email and has_contact:
                # Transform API response to match our Excel format
                transformed_applicant = {
                    "Date": applicant.get("submitted_at", ""),
                    "Name": applicant.get("full_name", ""),
                    "Email ID": applicant.get("email", ""),
                    "Contact Number": applicant.get("contact_number", ""),
                    "LinkedIn Profile": applicant.get("linkedin_profile", ""),
                    "Resume": applicant.get("resume_url", ""),
                    "Interested Position": applicant.get("interested_position", ""),
                    "Current Role": applicant.get("currentrole", ""),
                    "Current Organization": applicant.get("current_organisation", ""),
                    "Total Years of Experience": str(applicant.get("total_experience", "")) if applicant.get("total_experience") else "",
                    "Current Location": applicant.get("current_location", ""),
                    "Location Preference": applicant.get("location_preference", ""),
                    "Current CTC per Annum": str(applicant.get("current_ctc", "")) if applicant.get("current_ctc") else "",
                    "Expected CTC per Annum": str(applicant.get("expected_ctc", "")) if applicant.get("expected_ctc") else "",
                    "Notice Period": str(applicant.get("notice_period", "")) if applicant.get("notice_period") else "",
                    "In Notice": "Yes" if applicant.get("currently_noticeperiod") else "No",
                    "Immediate Joiner": "Yes" if applicant.get("immediate_joiner") else "No",
                    "Offers in Hand": "Yes" if applicant.get("other_offer_in_hand") else "No",
                    "Offered CTC": str(applicant.get("offered_ctc", "")) if applicant.get("offered_ctc") else "",
                    "Certifications": applicant.get("certifications", ""),
                    "Referred By": applicant.get("referred_by", ""),
                    "Interview Status": "",
                    "Application Status": "",
                    "Initial Screening": "",
                    "Round 1 D and T": "",
                    "Round 1 Remarks": "",
                    "Round 2 D and T": "",
                    "Round 2 Remarks": "",
                    "Offered Position": "",
                    "Joining Date": "",
                    "Reject Mail Sent": "No",
                    "Remarks": applicant.get("additional_info", ""),
                    "Screened By": ""
                }
                valid_applicants.append(transformed_applicant)
        
        logger.info(f"Filtered to {len(valid_applicants)} valid applicants (excluding null records)")
        logger.info("=== API fetch completed successfully ===")
        
        return jsonify({
            "status": "success",
            "data": valid_applicants,
            "total_count": len(raw_applicants),
            "valid_count": len(valid_applicants),
            "is_admin": is_admin()
        })
        
    except requests.exceptions.Timeout:
        logger.error("API request timed out")
        return jsonify({
            "status": "error",
            "message": "Request timed out. The dev server is slow, please try again."
        }), 504
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching applicants from API: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"Failed to fetch applicants: {str(e)}"
        }), 500
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"An unexpected error occurred: {str(e)}"
        }), 500

@app.route('/analytics')
@admin_required
def analytics():
    return render_template('analytics.html', is_admin=True)

@app.route('/api/analytics', methods=['GET'])
@login_required
def get_analytics_data():
    """Return analytics data for the dashboard"""
    try:
        # Get year parameter from query string
        year_filter = request.args.get('year')
        
        # Load data from Excel file
        data = load_data()
        
        # Filter data by year if specified
        if year_filter:
            filtered_data = []
            for item in data:
                date_str = item.get('Date of Application')
                if date_str:
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        if str(date_obj.year) == year_filter:
                            filtered_data.append(item)
                    except ValueError:
                        pass
            data = filtered_data
            print(f"Filtered data for year {year_filter}: {len(data)} records")
        else:
            print(f"No year filter applied, returning all {len(data)} records")
        
        # Calculate metrics
        total_applicant = len(data)
        total_rejected = sum(1 for item in data if item.get('Application Status') == 'Rejected')
        no_response = sum(1 for item in data if item.get('Application Status') == 'No Resp Call/Email')
        not_interviewed = sum(1 for item in data if item.get('Interview Status') == 'Not Interviewed')
        total_round_2_completed = sum(1 for item in data if item.get('Round 2 Status') == 'Completed')
        did_not_join = sum(1 for item in data if item.get('Application Status') == 'Did Not Join')
        on_hold = sum(1 for item in data if item.get('Application Status') == 'On Hold')
        accepted_waiting_reference = sum(1 for item in data if item.get('Application Status') == 'Accepted')
        total_in_notice_yet_to_join = sum(1 for item in data if item.get('Application Status') == 'In Notice')
        total_joined = sum(1 for item in data if item.get('Application Status') == 'Joined')
        intern = sum(1 for item in data if item.get('Interested Position') == 'Intern') # Assuming 'Intern' is a position

        # Monthly Statistics
        monthly_stats = defaultdict(lambda: {"applicants": 0, "accepted": 0, "rejected": 0, "in_notice": 0, "joined": 0})
        for item in data:
            date_str = item.get('Date of Application')
            if date_str:
                try:
                    # Assuming date format is 'YYYY-MM-DD' or similar
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    month_year = date_obj.strftime('%b %Y') # e.g., 'Nov 2025'
                    monthly_stats[month_year]["applicants"] += 1
                    if item.get('Application Status') == 'Accepted':
                        monthly_stats[month_year]["accepted"] += 1
                    elif item.get('Application Status') == 'Rejected':
                        monthly_stats[month_year]["rejected"] += 1
                    elif item.get('Application Status') == 'In Notice':
                        monthly_stats[month_year]["in_notice"] += 1
                    elif item.get('Application Status') == 'Joined':
                        monthly_stats[month_year]["joined"] += 1
                except ValueError:
                    # Handle cases where date_str might be in a different format or invalid
                    pass
        
        # Sort monthly statistics by date
        sorted_monthly_stats = []
        for month_year in sorted(monthly_stats.keys(), key=lambda x: datetime.strptime(x, '%b %Y')):
            stats = monthly_stats[month_year]
            sorted_monthly_stats.append({
                "month": month_year,
                "applicants": stats["applicants"],
                "accepted": stats["accepted"],
                "rejected": stats["rejected"],
                "in_notice": stats["in_notice"],
                "joined": stats["joined"]
            })

        # Hiring Funnel Status by Role
        hiring_funnel_by_role = defaultdict(lambda: {"count": 0})
        for item in data:
            role = item.get('Interested Position')
            if role:
                hiring_funnel_by_role[role]["count"] += 1
        
        sorted_hiring_funnel_by_role = [{"role": role, "count": stats["count"]} for role, stats in hiring_funnel_by_role.items()]

        # Position Statistics
        position_stats = defaultdict(lambda: {"applied": 0, "joined": 0})
        for item in data:
            position = item.get('Interested Position')
            if position:
                position_stats[position]["applied"] += 1
                if item.get('Application Status') == 'Joined':
                    position_stats[position]["joined"] += 1
        
        sorted_position_stats = [{"position": pos, "applied": stats["applied"], "joined": stats["joined"]} for pos, stats in position_stats.items()]
        
        # Offer Details - candidates with offered CTC
        offer_details = []
        for item in data:
            offered_ctc = item.get('Offered CTC')
            # Filter out invalid or non-offer values
            if offered_ctc and str(offered_ctc).lower() not in ['no', 'nil', 'n/a', '0', '', 'none']:
                offer_details.append({
                    'name': item.get('Name', 'N/A'),
                    'offered_ctc': offered_ctc,
                    'joining_date': item.get('Joining Date', 'N/A')
                })
        
        # User Activity - track users who have added screening/remarks (with monthly breakdown)
        user_activity = defaultdict(lambda: {"candidates_screened": 0})
        user_activity_by_month = defaultdict(lambda: defaultdict(lambda: {"candidates_screened": 0}))
        
        # Get current month for new screenings
        current_month = datetime.now().strftime('%b %Y')  # e.g., 'Jan 2026'
        
        for item in data:
            screened_by = item.get('Screened By')
            if screened_by and screened_by.strip():
                user_activity[screened_by]["candidates_screened"] += 1
                
                # Get date for monthly breakdown - use screening date if available, otherwise use application date
                # Priority: Screening Date > Initial Screening date parsing > Date > Date of Application
                date_str = None
                
                # Check for screening-related date fields
                initial_screening = item.get('Initial Screening')
                if initial_screening and initial_screening.strip():
                    # If there's initial screening content, try to extract date or use application date
                    date_str = item.get('Date') or item.get('Date of Application')
                else:
                    date_str = item.get('Date') or item.get('Date of Application')
                
                month_year = current_month  # Default to current month
                
                if date_str:
                    try:
                        date_str = str(date_str).strip()
                        # Try multiple date formats
                        if 'T' in date_str:
                            # ISO format from API: 2025-12-23T00:00:00.000Z
                            date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                        elif '/' in date_str:
                            # Excel format: M/D/YYYY or MM/DD/YYYY
                            date_obj = datetime.strptime(date_str.split(' ')[0], '%m/%d/%Y')
                        elif '-' in date_str:
                            # ISO date: YYYY-MM-DD
                            if ' ' in date_str:
                                date_obj = datetime.strptime(date_str.split(' ')[0], '%Y-%m-%d')
                            else:
                                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        else:
                            date_obj = None
                        
                        if date_obj:
                            month_year = date_obj.strftime('%b %Y')
                    except (ValueError, TypeError) as e:
                        logger.debug(f"Date parsing error for '{date_str}': {e}")
                        month_year = current_month  # Fallback to current month
                
                user_activity_by_month[month_year][screened_by]["candidates_screened"] += 1
        
        sorted_user_activity = [
            {"username": username, "candidates_screened": stats["candidates_screened"]} 
            for username, stats in sorted(user_activity.items(), key=lambda x: x[1]["candidates_screened"], reverse=True)
        ]
        
        # Format monthly user activity
        monthly_user_activity = {}
        for month_year, users in user_activity_by_month.items():
            monthly_user_activity[month_year] = [
                {"username": username, "candidates_screened": stats["candidates_screened"]}
                for username, stats in sorted(users.items(), key=lambda x: x[1]["candidates_screened"], reverse=True)
            ]
        
        return jsonify({
            'total_applicant': total_applicant,
            'total_rejected': total_rejected,
            'no_response': no_response,
            'not_interviewed': not_interviewed,
            'total_round_2_completed': total_round_2_completed,
            'did_not_join': did_not_join,
            'on_hold': on_hold,
            'accepted_waiting_reference': accepted_waiting_reference,
            'total_in_notice_yet_to_join': total_in_notice_yet_to_join,
            'total_joined': total_joined,
            'intern': intern,
            'monthly_statistics': sorted_monthly_stats,
            'hiring_funnel_by_role': sorted_hiring_funnel_by_role,
            'position_statistics': sorted_position_stats,
            'offer_details': offer_details,
            'user_activity': sorted_user_activity,
            'monthly_user_activity': monthly_user_activity
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/dropdown-options', methods=['GET'])
@login_required
def get_dropdown_options():
    """Return all dropdown options for form fields"""
    dropdown_options = {
        'Interested Position': [
            'Site Reliability Engineer',
            'Senior Site Reliability Engineer',
            'Lead Site Reliability Engineer',
            'Application Site Reliability Engineer',
            'Security Operations Centre Engineer',
            'Performance Engineer',
            'QA Automation Engineer (Playwright & Selenium)',
            'DevOps Engineer',
            'Lead SAP Engineer',
            'AI/ML Engineer',
            'AI/ML Intern',
            'Internship',
            'Fresher'
        ],
        'Current Role': [
            'Software Engineer',
            'Senior Software Engineer',
            'Lead Engineer',
            'Engineering Manager',
            'Architect',
            'QA Engineer',
            'DevOps Engineer',
            'Data Engineer',
            'Data Scientist',
            'Product Manager',
            'UI/UX Designer'
        ],
        'Current Location': [
            'Bangalore', 
            'Chennai', 
            'Coimbatore', 
            'Others'
        ],
        'Location Preference': [
            'Bangalore', 
            'Chennai', 
            'Coimbatore', 
            'Others'
        ],
        'Total Years of Experience': [
            '0-1 years',
            '1-2 years',
            '2-3 years',
        ],
        'Notice Period': [
            'Immediate',
            '15 days',
            '30 days',
            '60 days',
            '90 days'
        ],
        'In Notice': ['Yes', 'No'],
        'Immediate Joiner': ['Yes', 'No'],
        'Offers in Hand': ['Yes', 'No'],
        'Interview Status': [
            'Applied',
            'Profile Screening Comp',
            'Voice Screening Comp',
            'Tech Inter Sched',
            'Tech Inter Comp',
            'Code Inter Sched',
            'Code Inter Comp',
            'HR Inter Sched',
            'HR Inter Comp',
            'Offer',
            'Pending Final Noti',
            'References',
            'All Completed'
        ],
        'Application Status': [
            'Proceed Further',
            'On Hold',
            'No Resp Call/Email',
            'Did Not Join',
            'Sent',
            'Recieved',
            'In Notice',
            'Accepted',
            'Rejected',
            'Joined'
        ],
        'Reject Mail Sent': ['Yes', 'No']
    }
    
    return jsonify(dropdown_options)

# User Management Routes (Admin Only)
@app.route('/users')
@admin_required
def users_page():
    """User management page (admin only)"""
    return render_template('users.html', isAdmin=True)

@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    """Get all users (admin only)"""
    conn = sqlite3.connect(USER_DB)
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, is_admin, created_at FROM users ORDER BY created_at DESC')
    users = cursor.fetchall()
    conn.close()
    
    users_list = []
    for user in users:
        users_list.append({
            'id': user[0],
            'username': user[1],
            'is_admin': bool(user[2]),
            'created_at': user[3]
        })
    
    return jsonify(users_list)

@app.route('/api/users', methods=['POST'])
@admin_required
def add_user():
    """Add a new user (admin only)"""
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')
        is_admin_flag = data.get('is_admin', False)
        
        if not username or not password:
            return jsonify({"status": "error", "message": "Username and password are required"}), 400
        
        conn = sqlite3.connect(USER_DB)
        cursor = conn.cursor()
        
        # Check if username already exists
        cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', (username,))
        if cursor.fetchone()[0] > 0:
            conn.close()
            return jsonify({"status": "error", "message": "Username already exists"}), 400
        
        # Add user
        password_hash = hash_password(password)
        cursor.execute('''
            INSERT INTO users (username, password_hash, is_admin)
            VALUES (?, ?, ?)
        ''', (username, password_hash, 1 if is_admin_flag else 0))
        conn.commit()
        conn.close()
        
        return jsonify({"status": "success", "message": "User added successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    """Delete a user (admin only)"""
    try:
        # Prevent deleting yourself
        if user_id == session.get('user_id'):
            return jsonify({"status": "error", "message": "Cannot delete your own account"}), 400
        
        conn = sqlite3.connect(USER_DB)
        cursor = conn.cursor()
        
        # Check if user exists
        cursor.execute('SELECT username FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        if not user:
            conn.close()
            return jsonify({"status": "error", "message": "User not found"}), 404
        
        # Prevent deleting admin user
        cursor.execute('SELECT is_admin FROM users WHERE id = ?', (user_id,))
        if cursor.fetchone()[0] == 1:
            # Check if there are other admins
            cursor.execute('SELECT COUNT(*) FROM users WHERE is_admin = 1 AND id != ?', (user_id,))
            if cursor.fetchone()[0] == 0:
                conn.close()
                return jsonify({"status": "error", "message": "Cannot delete the last admin user"}), 400
        
        # Delete user
        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        conn.close()
        
        return jsonify({"status": "success", "message": "User deleted successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/send-rejection-email', methods=['POST'])
@login_required
def send_rejection_email_api():
    """API endpoint to send rejection email"""
    try:
        data = request.json
        candidate_name = data.get('name', 'Candidate')
        candidate_email = data.get('email')
        position = data.get('position', 'the position')
        
        if not candidate_email:
            return jsonify({
                "status": "error",
                "message": "Candidate email is required"
            }), 400
        
        # Send the email
        success, message = send_rejection_email(candidate_name, candidate_email, position)
        
        if success:
            return jsonify({
                "status": "success",
                "message": "Rejection email sent successfully"
            })
        else:
            return jsonify({
                "status": "error",
                "message": f"Failed to send email: {message}"
            }), 500
    
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

# Resume Filter Routes (Admin Only)
@app.route('/resume-filter')
@admin_required
def resume_filter_page():
    """Resume filter page (admin only)"""
    return render_template('resume_filter.html', is_admin=True)


@app.route('/api/resume-filter', methods=['POST'])
@admin_required
def filter_resumes():
    """API endpoint to filter resumes by keyword"""
    try:
        data = request.json
        keyword = data.get('keyword', '').strip().lower()
        
        if not keyword:
            return jsonify({
                "status": "error",
                "message": "Keyword is required"
            }), 400
        
        # Split keywords by space or comma and filter out empty strings
        import re
        keyword_terms = [t.strip().lower() for t in re.split(r'[,\s]+', keyword) if t.strip()]
        
        if not keyword_terms:
            return jsonify({
                "status": "error",
                "message": "Valid keywords are required"
            }), 400
        
        # Load candidate data
        candidates = load_data()
        results = []
        
        for index, candidate in enumerate(candidates):
            resume_filename = candidate.get('Resume', '')
            
            # Skip if no resume file
            if not resume_filename or resume_filename.startswith('http'):
                continue
            
            # Build full path to resume
            resume_path = os.path.join(app.config['UPLOAD_FOLDER'], resume_filename)
            
            # Skip if file doesn't exist
            if not os.path.exists(resume_path):
                continue
            
            # Extract text from PDF
            try:
                reader = PdfReader(resume_path)
                resume_text = ""
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        resume_text += page_text.lower()
                
                # Count occurrences for each term and sum them up
                total_matches = 0
                term_matches = {}
                for term in keyword_terms:
                    count = resume_text.count(term)
                    if count > 0:
                        total_matches += count
                        term_matches[term] = count
                
                if total_matches > 0:
                    results.append({
                        'index': index,
                        'name': candidate.get('Name', 'N/A'),
                        'email': candidate.get('Email ID', 'N/A'),
                        'position': candidate.get('Interested Position', 'N/A'),
                        'experience': candidate.get('Total Years of Experience', 'N/A'),
                        'match_score': total_matches,
                        'term_matches': term_matches,
                        'resume': resume_filename
                    })
            except Exception as pdf_error:
                logger.warning(f"Error reading PDF {resume_filename}: {pdf_error}")
                continue
        
        # Sort by match score (highest first)
        results.sort(key=lambda x: x['match_score'], reverse=True)
        
        return jsonify({
            "status": "success",
            "keyword": keyword,
            "total_matches": len(results),
            "results": results
        })
    
    except Exception as e:
        logger.error(f"Error in resume filter: {e}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


if __name__ == '__main__':
    init_user_db()
    # Apply header ordering to existing Excel data on startup
    try:
        data = load_data()
        save_data(data)
    except Exception as e:
        print(f"Error applying header ordering on startup: {e}")
    app.run(debug=app.config.get("DEBUG", False), port=app.config.get("PORT", 5000))
